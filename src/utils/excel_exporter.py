"""
Chunked Excel / CSV exporter for large datasets (100 k+ rows).

* **XLSX** – uses *xlsxwriter* in streaming mode (constant memory).
* **CSV**  – uses the stdlib *csv* module with server-side cursor chunks.

Both modes fetch data from the database in configurable chunks so memory
stays bounded even for 300 k+ row result sets.
"""
from __future__ import annotations

import csv
import logging
import math
import os
from datetime import date, datetime
from typing import Dict, Iterator, List, Optional, Tuple

import pandas as pd
import xlsxwriter
from psycopg2.extras import execute_values

logger = logging.getLogger(__name__)

CHUNK_SIZE = 50_000            # rows fetched per round-trip
EXCEL_MAX_ROWS = 1_048_576     # Excel hard limit per sheet


# ---------------------------------------------------------------------------
# Internal: chunked DB fetch
# ---------------------------------------------------------------------------

def _fetch_chunks(
    conn,
    query: str,
    chunk_size: int = CHUNK_SIZE,
) -> Iterator[pd.DataFrame]:
    """Execute *query* and yield DataFrames of *chunk_size* rows."""
    cur = conn.cursor()
    try:
        cur.execute(query)
        cols = [d[0] for d in cur.description]
        while True:
            rows = cur.fetchmany(chunk_size)
            if not rows:
                break
            yield pd.DataFrame(rows, columns=cols)
    finally:
        cur.close()


# ---------------------------------------------------------------------------
# Data-preparation helpers (staging INSERT / TRUNCATE)
# ---------------------------------------------------------------------------

def truncate_staging_table(conn, schema: str, table: str) -> None:
    """TRUNCATE a staging table. Commits immediately."""
    cur = conn.cursor()
    try:
        cur.execute(f"TRUNCATE TABLE {schema}.{table}")
        conn.commit()
        logger.info(f"[PREP] TRUNCATED {schema}.{table}")
    finally:
        cur.close()


def fetch_prep_data(conn, query: str) -> Tuple[List[str], List[Tuple]]:
    """
    Execute a SELECT query and return (columns, rows) as plain Python objects.

    Uses a generator-based fetch (per boilerplate README) and converts to
    a list of tuples suitable for ``execute_values``.

    Commits after fetch to fully release the DB transaction/cursor,
    preventing overlapping active queries in PostgreSQL when called
    inside a loop.

    Returns
    -------
    columns : list[str]
        Column names from the query result.
    rows : list[tuple]
        All result rows as tuples.
    """
    cur = conn.cursor()
    try:
        cur.execute(query)
        columns = [d[0] for d in cur.description]
        all_rows: List[Tuple] = []
        while True:
            batch = cur.fetchmany(CHUNK_SIZE)
            if not batch:
                break
            all_rows.extend(batch)
        logger.info(f"[PREP] Fetched {len(all_rows)} rows ({len(columns)} columns)")
        return columns, all_rows
    finally:
        cur.close()
        conn.commit()


def insert_prep_data(
    conn,
    schema: str,
    table: str,
    columns: List[str],
    data: List[Tuple],
    page_size: int = 10_000,
) -> int:
    """
    Bulk INSERT into a staging table using ``psycopg2.extras.execute_values``.

    Follows the **Insert Multiple Values** pattern recommended by the
    boilerplate README.  Data is sent in pages of *page_size* rows
    (default 10 000) for efficient network utilisation.

    Parameters
    ----------
    conn
        A *psycopg2* connection (autocommit=False assumed).
    schema : str
        Target schema (e.g. ``'dad_dev'``).
    table : str
        Target table  (e.g. ``'temp_week_robotic_summary'``).
    columns : list[str]
        Column names matching the order of values in *data*.
    data : list[tuple]
        Rows to insert – one tuple per row.
    page_size : int
        Rows per VALUES sub-batch (default 10 000).

    Returns
    -------
    int – total rows inserted.
    """
    if not data:
        logger.info(f"[PREP] No data to insert into {schema}.{table}")
        return 0

    cur = conn.cursor()
    col_str = ", ".join(columns)
    sql = f"INSERT INTO {schema}.{table} ({col_str}) VALUES %s"

    try:
        execute_values(cur, sql, data, page_size=page_size)
        conn.commit()
        logger.info(
            f"[PREP] Inserted {len(data)} rows into {schema}.{table} "
            f"(page_size={page_size})"
        )
        return len(data)
    except Exception:
        conn.rollback()
        logger.error(f"[PREP] INSERT failed for {schema}.{table}")
        raise
    finally:
        cur.close()


# ---------------------------------------------------------------------------
# XLSX export (XlsxWriter – streaming, write-once)
# ---------------------------------------------------------------------------

def export_xlsx(
    output_path: str,
    sheets_config: List[Tuple[str, str]],
    conn,
    format_config: Optional[Dict] = None,
    chunk_size: int = CHUNK_SIZE,
) -> str:
    """
    Export multiple queries to a multi-sheet ``.xlsx`` workbook.

    Each tuple in *sheets_config* is ``(sheet_name, sql_query)``.
    If a single sheet would exceed ~1 M rows the data is automatically
    split across numbered continuation sheets.

    Parameters
    ----------
    output_path : str
        Destination ``.xlsx`` file (created / overwritten).
    sheets_config : list[tuple[str, str]]
        ``[(sheet_name, sql_query), ...]``
    conn
        A *psycopg2* connection.
    format_config : dict, optional
        ``header_bold`` (bool), ``auto_width`` (bool),
        ``header_bg_color`` (str hex).
    chunk_size : int
        Rows per DB round-trip (default 50 000).

    Returns
    -------
    str  – the absolute *output_path* written.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    fmt = format_config or {}
    workbook = xlsxwriter.Workbook(output_path, {"constant_memory": True, "use_zip64": True})

    # ---- formats ---------------------------------------------------------
    hdr_fmt = workbook.add_format({
        "bold":     fmt.get("header_bold", True),
        "bg_color": fmt.get("header_bg_color", "#D3D3D3"),
        "border":   1,
        "align":    "center",
        "valign":   "vcenter",
    })
    txt_fmt = workbook.add_format({"border": 1, "valign": "vcenter"})
    num_fmt = workbook.add_format({"border": 1, "num_format": "#,##0.00", "align": "right"})
    date_fmt = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd", "align": "center"})
    dt_fmt = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd hh:mm:ss", "align": "center"})

    try:
        for sheet_name, sql in sheets_config:
            logger.info(f"[XLSX] Processing sheet: {sheet_name}")

            sheet_part = 0
            row = 0
            ws: xlsxwriter.worksheet.Worksheet | None = None
            headers_written = False
            columns: list = []
            col_fmts: dict = {}
            col_widths: dict = {}

            for chunk in _fetch_chunks(conn, sql, chunk_size):
                # first chunk → detect column formats once
                if not headers_written:
                    columns = list(chunk.columns)
                    for ci, col_name in enumerate(columns):
                        dtype = chunk.iloc[:, ci].dtype
                        if pd.api.types.is_datetime64_any_dtype(dtype):
                            col_fmts[ci] = dt_fmt
                        elif pd.api.types.is_float_dtype(dtype) or pd.api.types.is_integer_dtype(dtype):
                            col_fmts[ci] = num_fmt
                        else:
                            col_fmts[ci] = txt_fmt
                        col_widths[ci] = min(50, max(12, len(str(col_name)) + 2))

                # Do we need a new sheet (first time, or >1 M rows)?
                if ws is None or row >= EXCEL_MAX_ROWS:
                    suffix = f" ({sheet_part + 1})" if sheet_part > 0 else ""
                    ws_name = (sheet_name + suffix)[:31]
                    ws = workbook.add_worksheet(ws_name)
                    ws.freeze_panes(1, 0)
                    row = 0
                    sheet_part += 1

                    # write header row
                    for ci, col_name in enumerate(columns):
                        ws.write(row, ci, col_name, hdr_fmt)
                    row += 1
                    headers_written = True

                # write data rows
                for _, series in chunk.iterrows():
                    if row >= EXCEL_MAX_ROWS:
                        # spill into continuation sheet
                        suffix = f" ({sheet_part + 1})"
                        ws_name = (sheet_name + suffix)[:31]
                        ws = workbook.add_worksheet(ws_name)
                        ws.freeze_panes(1, 0)
                        row = 0
                        sheet_part += 1
                        for ci, col_name in enumerate(columns):
                            ws.write(row, ci, col_name, hdr_fmt)
                        row += 1

                    for ci, val in enumerate(series):
                        if val is None or (isinstance(val, float) and math.isnan(val)) or pd.isna(val):
                            ws.write_blank(row, ci, None, col_fmts.get(ci, txt_fmt))
                        elif isinstance(val, datetime):
                            ws.write_datetime(row, ci, val, dt_fmt)
                        elif isinstance(val, date):
                            ws.write_datetime(row, ci, datetime(val.year, val.month, val.day), date_fmt)
                        else:
                            ws.write(row, ci, val, col_fmts.get(ci, txt_fmt))
                    row += 1

                logger.info(f"  chunk written – sheet total so far: {row - 1} data rows")

            # auto-width
            if fmt.get("auto_width") and ws is not None:
                for ci, w in col_widths.items():
                    ws.set_column(ci, ci, w)

            if row <= 1:
                logger.warning(f"Sheet '{sheet_name}': 0 data rows (headers only)")

            logger.info(f"[XLSX] Sheet '{sheet_name}' done – {row - 1} data rows, {sheet_part} part(s)")

    finally:
        workbook.close()

    logger.info(f"[XLSX] Saved: {output_path}")
    return os.path.abspath(output_path)


def export_dataframe_xlsx(
    output_path: str,
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    format_config: Optional[Dict] = None,
    chunk_size: int = CHUNK_SIZE,
) -> int:
    """
    Export a single DataFrame to ``.xlsx`` using XlsxWriter in
    ``constant_memory`` mode with row-wise writes.

    This is the DataFrame counterpart of :func:`export_xlsx` (which
    streams from a SQL query).  Use this when data has already been
    loaded into a DataFrame (e.g. after a cross-DB merge).

    The DataFrame is iterated in chunks of *chunk_size* rows so that
    the internal write buffer stays bounded.  Each chunk is written
    and then dereferenced before the next one is processed.

    Parameters
    ----------
    output_path : str
        Destination ``.xlsx`` file (created / overwritten).
    df : pd.DataFrame
        The data to export.
    sheet_name : str
        Worksheet name (truncated to 31 chars per Excel limit).
    format_config : dict, optional
        ``header_bold`` (bool), ``auto_width`` (bool),
        ``header_bg_color`` (str hex).
    chunk_size : int
        Rows written per iteration (default 50 000).

    Returns
    -------
    int – total data rows written.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    fmt = format_config or {}
    workbook = xlsxwriter.Workbook(output_path, {"constant_memory": True, "use_zip64": True})

    # ---- formats ---------------------------------------------------------
    hdr_fmt = workbook.add_format({
        "bold":     fmt.get("header_bold", True),
        "bg_color": fmt.get("header_bg_color", "#D3D3D3"),
        "border":   1,
        "align":    "center",
        "valign":   "vcenter",
    })
    txt_fmt = workbook.add_format({"border": 1, "valign": "vcenter"})
    num_fmt = workbook.add_format({"border": 1, "num_format": "#,##0.00", "align": "right"})
    date_fmt = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd", "align": "center"})
    dt_fmt = workbook.add_format({"border": 1, "num_format": "yyyy-mm-dd hh:mm:ss", "align": "center"})

    columns = list(df.columns)

    # detect per-column format
    col_fmts: dict = {}
    for ci, col_name in enumerate(columns):
        dtype = df.iloc[:, ci].dtype
        if pd.api.types.is_datetime64_any_dtype(dtype):
            col_fmts[ci] = dt_fmt
        elif pd.api.types.is_float_dtype(dtype) or pd.api.types.is_integer_dtype(dtype):
            col_fmts[ci] = num_fmt
        else:
            col_fmts[ci] = txt_fmt

    total_rows = len(df)
    n_chunks = math.ceil(total_rows / chunk_size) if total_rows else 0

    try:
        sheet_part = 0
        row = 0
        ws = workbook.add_worksheet(sheet_name[:31])
        ws.freeze_panes(1, 0)
        sheet_part += 1

        # write header
        for ci, col_name in enumerate(columns):
            ws.write(row, ci, col_name, hdr_fmt)
        row += 1

        # write data in chunks
        for chunk_idx in range(n_chunks):
            start = chunk_idx * chunk_size
            end = min(start + chunk_size, total_rows)
            chunk = df.iloc[start:end]

            for _, series in chunk.iterrows():
                # spill into continuation sheet if >1 M rows
                if row >= EXCEL_MAX_ROWS:
                    suffix = f" ({sheet_part + 1})"
                    ws_name = (sheet_name + suffix)[:31]
                    ws = workbook.add_worksheet(ws_name)
                    ws.freeze_panes(1, 0)
                    row = 0
                    sheet_part += 1
                    for ci, col_name in enumerate(columns):
                        ws.write(row, ci, col_name, hdr_fmt)
                    row += 1

                for ci, val in enumerate(series):
                    if val is None or (isinstance(val, float) and math.isnan(val)) or pd.isna(val):
                        ws.write_blank(row, ci, None, col_fmts.get(ci, txt_fmt))
                    elif isinstance(val, datetime):
                        ws.write_datetime(row, ci, val, dt_fmt)
                    elif isinstance(val, date):
                        ws.write_datetime(row, ci, datetime(val.year, val.month, val.day), date_fmt)
                    else:
                        ws.write(row, ci, val, col_fmts.get(ci, txt_fmt))
                row += 1

            logger.info(f"  chunk {chunk_idx + 1}/{n_chunks} written – {row - 1} data rows so far")

    finally:
        workbook.close()

    logger.info(f"[XLSX-DF] Saved: {output_path} – {row - 1} data rows, {sheet_part} part(s)")
    return row - 1


# ---------------------------------------------------------------------------
# Append sheets to an existing workbook (openpyxl – loads file into memory)
# ---------------------------------------------------------------------------

def append_sheets_to_xlsx(
    xlsx_path: str,
    sheets_config: List[Tuple[str, str]],
    conn,
    chunk_size: int = CHUNK_SIZE,
) -> str:
    """
    Open an **existing** ``.xlsx`` and add new sheets without losing
    previously written data.

    Use this when the first batch of sheets has already been created with
    :func:`export_xlsx` and subsequent queries must be appended to the
    same file.

    .. warning::
        *openpyxl* loads the whole workbook into memory.  Avoid calling this
        on files larger than ~500 MB; prefer creating all sheets in one
        :func:`export_xlsx` call when possible.
    """
    from openpyxl import load_workbook

    logger.info(f"[XLSX-APPEND] Opening {xlsx_path}")
    wb = load_workbook(xlsx_path)

    for sheet_name, sql in sheets_config:
        ws = wb.create_sheet(sheet_name[:31])
        row_num = 1
        headers_written = False

        for chunk in _fetch_chunks(conn, sql, chunk_size):
            if not headers_written:
                for ci, col_name in enumerate(chunk.columns, 1):
                    ws.cell(row_num, ci, col_name)
                row_num += 1
                headers_written = True

            for _, series in chunk.iterrows():
                for ci, val in enumerate(series, 1):
                    ws.cell(row_num, ci, val)
                row_num += 1

        logger.info(f"[XLSX-APPEND] Sheet '{sheet_name}' – {row_num - 2} data rows")

    wb.save(xlsx_path)
    logger.info(f"[XLSX-APPEND] Saved: {xlsx_path}")
    return os.path.abspath(xlsx_path)


# ---------------------------------------------------------------------------
# CSV export (stdlib csv – most memory-efficient)
# ---------------------------------------------------------------------------

def export_csv(
    output_path: str,
    sql_query: str,
    conn,
    chunk_size: int = CHUNK_SIZE,
) -> str:
    """
    Stream query results straight to a ``.csv`` file.

    Memory usage is constant (~10 MB) regardless of total row count.

    Returns the absolute path of the written file.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    cur = conn.cursor()
    try:
        cur.execute(sql_query)
        columns = [d[0] for d in cur.description]

        # utf-8-sig so Excel opens the file with correct encoding
        with open(output_path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.writer(fh, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(columns)

            total = 0
            while True:
                rows = cur.fetchmany(chunk_size)
                if not rows:
                    break
                writer.writerows(rows)
                total += len(rows)
                logger.info(f"[CSV] chunk {math.ceil(total / chunk_size)}: "
                            f"{len(rows)} rows (total {total})")

        logger.info(f"[CSV] Saved: {output_path} – {total} rows")
    finally:
        cur.close()

    return os.path.abspath(output_path)


# ---------------------------------------------------------------------------
# Template-based export (openpyxl – load template, inject data)
# ---------------------------------------------------------------------------

def update_template_with_data(
    template_path: str,
    output_path: str,
    sheets_config: List[Tuple[str, str]],
    conn,
    header_rows: int = 1,
    chunk_size: int = CHUNK_SIZE,
    format_config: Optional[Dict] = None,
) -> str:
    """
    Load an existing Excel template, inject query data into named sheets,
    and save to *output_path*.

    The template is **never modified** -- a copy is written to *output_path*.

    For each ``(sheet_name, sql)`` pair in *sheets_config*:
      - Find the sheet in the template by *sheet_name*.
      - Skip the first *header_rows* rows (preserve template headers).
      - Append data starting at row ``header_rows + 1``.

    If a sheet_name does not exist in the template, a new sheet is created.

    Args:
        template_path: path to the .xlsx template file.
        output_path:   path for the output file (copy of template + data).
        sheets_config: list of (sheet_name, sql_query) tuples.
        conn:          database connection.
        header_rows:   number of header rows in template to preserve.
        chunk_size:    rows per DB fetch.
        format_config: unused for template mode (formatting comes from template).

    Returns:
        Absolute path of the generated file.
    """
    import shutil
    from openpyxl import load_workbook

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Copy template to output location (never modify original)
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    shutil.copy2(template_path, output_path)
    logger.info(f"[TEMPLATE] Copied template -> {output_path}")

    wb = load_workbook(output_path)
    existing_sheets = wb.sheetnames

    for sheet_name, sql in sheets_config:
        ws_name = sheet_name[:31]

        if ws_name in existing_sheets:
            ws = wb[ws_name]
            logger.info(f"[TEMPLATE] Writing data to existing sheet '{ws_name}' "
                        f"(start row {header_rows + 1})")
        else:
            ws = wb.create_sheet(ws_name)
            header_rows_for_sheet = 0  # new sheet has no headers
            logger.info(f"[TEMPLATE] Created new sheet '{ws_name}'")

        # Determine start row: after header rows for existing sheets
        start_row = header_rows + 1 if ws_name in existing_sheets else 1

        # Clear any old data below header rows (in case of re-run)
        if ws_name in existing_sheets and ws.max_row >= start_row:
            for row_cells in ws.iter_rows(min_row=start_row,
                                          max_row=ws.max_row,
                                          max_col=ws.max_column):
                for cell in row_cells:
                    cell.value = None

        current_row = start_row
        total_rows = 0

        for chunk in _fetch_chunks(conn, sql, chunk_size):
            # For new sheets (not in template), write headers from first chunk
            if ws_name not in existing_sheets and current_row == 1:
                for ci, col_name in enumerate(chunk.columns, 1):
                    ws.cell(current_row, ci, col_name)
                current_row += 1

            for _, series in chunk.iterrows():
                for ci, val in enumerate(series, 1):
                    cell = ws.cell(current_row, ci)
                    if isinstance(val, datetime):
                        cell.value = val
                        cell.number_format = "YYYY-MM-DD HH:MM:SS"
                    elif isinstance(val, date):
                        cell.value = val
                        cell.number_format = "YYYY-MM-DD"
                    else:
                        cell.value = val
                current_row += 1
                total_rows += 1

        logger.info(f"[TEMPLATE] Sheet '{ws_name}' done - {total_rows} data rows")

    wb.save(output_path)
    logger.info(f"[TEMPLATE] Saved: {output_path}")
    return os.path.abspath(output_path)


# ---------------------------------------------------------------------------
# High-level convenience wrapper
# ---------------------------------------------------------------------------

def export_report(
    output_path: str,
    sheets_config: List[Tuple[str, str]],
    conn,
    output_format: str = "xlsx",
    append: bool = False,
    format_config: Optional[Dict] = None,
    chunk_size: int = CHUNK_SIZE,
    template_file: Optional[str] = None,
    template_header_rows: int = 1,
) -> str:
    """
    One-call entry point used by the report-generator job.

    * If *template_file* is provided and exists -> load template, inject data.
    * ``output_format='xlsx'``  -> multi-sheet Excel via XlsxWriter.
    * ``output_format='csv'``   -> one CSV per query (sheet_name used in filename).

    When *append* is ``True`` **and** the file already exists, new sheets are
    added to the existing workbook (openpyxl path).

    Returns the path (or directory for CSV) of the generated output.
    """
    output_format = output_format.lower().strip()

    # Template mode: load template, inject data, save
    if template_file and os.path.exists(template_file):
        logger.info(f"[TEMPLATE] Using template: {template_file}")
        return update_template_with_data(
            template_path=template_file,
            output_path=output_path,
            sheets_config=sheets_config,
            conn=conn,
            header_rows=template_header_rows,
            chunk_size=chunk_size,
            format_config=format_config,
        )

    # No template -> auto-generate
    if output_format == "xlsx":
        if append and os.path.exists(output_path):
            return append_sheets_to_xlsx(output_path, sheets_config, conn, chunk_size)
        return export_xlsx(output_path, sheets_config, conn, format_config, chunk_size)

    if output_format == "csv":
        # CSV: one file per query, stored next to each other
        base_dir = os.path.dirname(output_path) or "."
        base_name = os.path.splitext(os.path.basename(output_path))[0]
        os.makedirs(base_dir, exist_ok=True)

        paths: list[str] = []
        for sheet_name, sql_query in sheets_config:
            safe_name = sheet_name.replace(" ", "_").replace("/", "_")
            csv_path = os.path.join(base_dir, f"{base_name}_{safe_name}.csv")
            export_csv(csv_path, sql_query, conn, chunk_size)
            paths.append(csv_path)

        return base_dir  # return directory containing all CSV files

    raise ValueError(f"Unsupported output_format: {output_format}. Use 'xlsx' or 'csv'.")
